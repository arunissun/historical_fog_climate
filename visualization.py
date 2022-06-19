#!/usr/bin/env python
# coding: utf-8

# ### Here we present the code for visualizing the figures which are embeded in the paper 

# ### This code was applied for all the years and stations (Zagreb, Budapest, Hereny, Nagyszeben, Pecs, Ungvar)
# ### Here we show an example of Zagreb

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import glob
from openpyxl import load_workbook
## reading an excel file containing data by specifying path (year and station name)
def read_data(path,name = str):
    
    df = pd.read_excel(path, sheet_name= name)
    return df

## Here we list all the data files (excel files) 
## in which 'Zagreb' sheet is present
## The code snippet mentioned below can be used for any other station also
## Just change the station name in the code
def files():
    files_1900_1919 = []
    files_1886_1899 = []
    
    for name in glob.glob('/Users/varungandhi/Downloads/Research_paper/paper1/data_paper/Data/1900 - 1919/*'):
        files_1900_1919.append(name)
        
    for name in glob.glob('/Users/varungandhi/Downloads/Research_paper/paper1/data_paper/Data/1880 - 1899/*'):
        files_1886_1899.append(name)
        
    all_files = files_1886_1899+files_1900_1919
    return all_files
        
    
##example for reading the files for 'Zagreb'

def read_data():
    full_data = files()
    station_data = []
    
    for path_name in full_data:
        wb = load_workbook(path_name, read_only=True) # open an Excel file and return a workbook
        if 'Zagreb' in wb.sheetnames:
            df = pd.read_excel(path_name,sheet_name = 'Zagreb')
            station_data.append(df)
        
    return station_data


# creating 12 dataframes, one data frame for each month by 
# specifying the row and column ranges for 
# the particular month
# Here we present the example for the station - 'Zagreb', 
# similar dataframes for 
# every year and other stations
# can be produced

def create_df(ra,rb,rc,rd,raa,rbb,rcc,rdd,c1,c2,c3,c4,c5,c6,data):
    df1 =  data.iloc[ra:raa,  c1:c2]
    df2 =  data.iloc[ra:raa-3,c3:c4]
    df3 =  data.iloc[ra:raa,  c5:c6]
    df4 =  data.iloc[rb:rbb,  c1:c2]
    df5 =  data.iloc[rb:rbb+1,c3:c4]
    df6 =  data.iloc[rb:rbb,  c5:c6]
    df7 =  data.iloc[rc:rcc,  c1:c2]
    df8 =  data.iloc[rc:rcc,  c3:c4]
    df9 =  data.iloc[rc:rcc-1,c5:c6]
    df10 = data.iloc[rd:rdd,  c1:c2]
    df11 = data.iloc[rd:rdd-1,c3:c4]
    df12 = data.iloc[rd:rdd,  c5:c6]
    return [df1,df2,df3,df4,df5,df6,df7,df8,df9,df10,df11,df12]

## the yearly_data function will
## create monthly dataframes for all the years (1886 - 1914 for Zagreb)
def yearly_data():
    station_data = read_data()
    monthwise = []
    for years in station_data:
        months_datalist = create_df(ra = 3,rb = 52,rc = 97,rd = 141,raa = 34,rbb = 82,rcc = 128,
                   rdd = 172,c1 = 2,c2 = 5,c3 = 9,c4 = 12,c5 = 15,c6 = 18,data = years)
        monthwise.append(months_datalist)
        
    return monthwise 

## The function monthly_data will enaable to 
## have seperate data frames for every month and in a specific year
def monthly_data():
    y_data = yearly_data()
    yearly_dataframes = []
    for years in range(len(y_data)):
    
        df1 =  y_data[years][0]
        df2 =  y_data[years][1]
        df3 =  y_data[years][2]
        df4 =  y_data[years][3]
        df5 =  y_data[years][4]
        df6 =  y_data[years][5]
        df7 =  y_data[years][6]
        df8 =  y_data[years][7]
        df9 =  y_data[years][8]
        df10 = y_data[years][9]
        df11 = y_data[years][10]
        df12 = y_data[years][11]
        yearly_dataframes = yearly_dataframes + [[df1,df2,df3,df4,df5,df6,df7,
                                    df8,df9,df10,df11,df12]]
        
    return yearly_dataframes

## The cal fucntion transforms 
## monthly data in to numpy arrays as well as omits any NAN values
## Finally the cal function will concatenate the monthly data into 
## yearly data
def cal():
    m_data = monthly_data()
    y_conct =[]
    for years in m_data:
        
        array_1 =  years[0].dropna().to_numpy(dtype = int)
        array_2 =  years[1].dropna().to_numpy(dtype= int)
        array_3 =  years[2].dropna().to_numpy(dtype= int)
        array_4 =  years[3].dropna().to_numpy(dtype= int)
        array_5 =  years[4].dropna().to_numpy(dtype= int)
        array_6 =  years[5].dropna().to_numpy(dtype= int)
        array_7 =  years[6].dropna().to_numpy(dtype= int)
        array_8 =  years[7].dropna().to_numpy(dtype= int)
        array_9 =  years[8].dropna().to_numpy(dtype= int)
        array_10 = years[9].dropna().to_numpy(dtype= int)
        array_11 = years[10].dropna().to_numpy(dtype= int)
        array_12 = years[11].dropna().to_numpy(dtype= int)
    
    #final_array = np.concatenate((array_1, array_2), axis = 0)
    
        final_array = np.concatenate((array_1, array_2, array_3,array_4, array_5,array_6,array_7, array_8, array_9, array_10, array_11, array_12), axis = 0)
        y_conct.append(final_array)
    return y_conct

y_all = cal()
years_total = np.arange(1886,1915,1)

## making a dictionary of data for Zagreb, where keys are the years and values are the data for that year
years_y = ['y_'+ str(x) for x in years_total]
dict_y = {x :y for  x,y in zip(years_y,y_all)}

## The cals function combines 
## data for all the years into single array
## and calculates consecutive fog events/year

def cals(arr1, arr2, arr3, arr4, arr5, arr6, arr7, arr8, arr9, arr10, arr11, arr12, arr13, arr14, arr15, arr16,
         arr17,arr18,arr19,arr20,arr21,arr22,arr23,arr24,arr25,arr26,arr27,
        arr28, arr29):
    
    years_array = np.concatenate((arr1, arr2, arr3, arr4, arr5, arr6, arr7,
                                  arr8, arr9, arr10, arr11, arr12, arr13, 
                                  arr14, arr15, arr16,arr17,arr18,arr19,
                                  arr20,arr21,arr22,arr23,arr24,arr25,
                                  arr26,arr27,arr28, arr29), axis = 0)
    final_array = years_array.reshape(years_array.shape[0]*years_array.shape[1])
    
    a_str = ''.join(str(x) for x in final_array)
    
    c_str = a_str.split('0')
    
    a1 = 0
    a2 = 0
    a3 = 0
    a4 = 0
    a5 = 0
    a6 = 0
    a7 = 0
    a8 = 0
    a9 = 0
    a10 = 0
    a11 = 0
    a12 = 0
    a13 = 0
    a14 = 0
    a15 = 0
    a16 = 0
    a17 = 0
    a18 = 0
    a19 = 0
    a20 = 0
    
    find_pattern = ['1', '11', '111', '1111','11111', '111111', '1111111', '11111111',
                   '111111111', '1111111111', '11111111111', '111111111111', '1111111111111',
                   '11111111111111','111111111111111', '1111111111111111', '11111111111111111',
                   '111111111111111111', '1111111111111111111', '11111111111111111111']

    for p in range(len(c_str)):
        if find_pattern[0] == c_str[p]:
            a1+=1
            
        if find_pattern[1] == c_str[p]:
            a2+=1
            
        if find_pattern[2] == c_str[p]:
            a3+=1
            
        if find_pattern[3] == c_str[p]:
            a4+=1
            
        if find_pattern[4] == c_str[p]:
            a5+=1
            
        if find_pattern[5] == c_str[p]:
            a6+=1
            
        if find_pattern[6] == c_str[p]:
            a7+=1
            
        if find_pattern[7] == c_str[p]:
            a8+=1
            
        if find_pattern[8] == c_str[p]:
            a9+=1
            
        if find_pattern[9] == c_str[p]:
            a10+=1
            
        if find_pattern[10] == c_str[p]:
            a11+=1
            
        if find_pattern[11] == c_str[p]:
            a12+=1
            
        if find_pattern[12] == c_str[p]:
            a13+=1
            
        if find_pattern[13] == c_str[p]:
            a14+=1
            
        if find_pattern[14] == c_str[p]:
            a15+=1
        
        if find_pattern[15] == c_str[p]:
            a16+=1
        
        if find_pattern[16] == c_str[p]:
            a17+=1
            
        if find_pattern[17] == c_str[p]:
            a18+=1
            
        if find_pattern[18] == c_str[p]:
            a19+=1
            
        if find_pattern[10] == c_str[p]:
            a20+=1
        
    return a1,a2,a3,a4,a5,a6,a7,a8,a9,a10,a11,a12,a13,a14,a15,a16,a17,a18,a19,a20


result = cals(dict_y['y_1886'],dict_y['y_1887'],dict_y['y_1888'],dict_y['y_1889'],
             dict_y['y_1890'],dict_y['y_1891'],dict_y['y_1892'],dict_y['y_1893'],
              dict_y['y_1894'],dict_y['y_1895'],
             dict_y['y_1896'],
             dict_y['y_1897'],
             dict_y['y_1898'],
             dict_y['y_1899'],
             dict_y['y_1900'],
             dict_y['y_1901'],
             dict_y['y_1902'],
             dict_y['y_1903'],
             dict_y['y_1904'],
             dict_y['y_1905'],
             dict_y['y_1906'],
             dict_y['y_1907'],
              dict_y['y_1908'],
              dict_y['y_1909'],
              dict_y['y_1910'],
              dict_y['y_1911'],
              dict_y['y_1912'],
              dict_y['y_1913'],
              dict_y['y_1914'])

# ### After running the above code for all the 6 stations (Budapest, Zagreb, Nagyszeben, Ungvar, Pecs, Hereny)
# ### we made data frames for visualizing figure 3a and 3b in the paper 

# In[80]:

print(result)
data = {'Budapest': [597, 114,  40,  23,  13,   8,   2,   4,   3,   3,   0,   0,   0,
                     0,   0,   0,   0,   0,   0,   0], 
        'Pécs': [502, 70, 50, 31, 13, 10, 8, 2, 4, 5, 3, 0, 1, 1, 2, 0, 0, 1, 0, 3],
        'Nagyszeben':[469, 96, 24, 21, 14, 6, 3, 1, 2, 2, 0, 1, 1, 1, 0, 0, 1, 0, 0, 0],
        'Ungvár':[361, 81, 43, 22, 16, 10, 4, 2, 0, 4, 0, 3, 1, 0, 0, 2, 0, 0, 0, 0],
        'Herény':[281, 64, 23, 13, 9, 8, 2, 6, 1, 1, 1, 1, 0, 0, 0, 0, 0, 0, 0, 1],
        'Zágráb':[579, 80, 31, 20, 7, 8, 4, 2, 0, 2, 0, 2, 0, 0, 2, 0, 0, 0, 0, 0]
       }


# In[81]:


## making the dataframe 
data = pd.DataFrame(data)
inde = ['1', '2', '3','4','5','6','7','8','9','10','11','12','13','14','15','16','17','18','19','20']
data.index = inde


# In[82]:


## Normalizing the data
data['Budapest'] = data['Budapest']/31
data['Pécs'] = data['Pécs']/27
data['Nagyszeben'] = data['Nagyszeben']/31
data['Ungvár'] = data['Ungvár']/28
data['Herény'] = data['Herény']/27
data['Zágráb'] = data['Zágráb']/29


# In[83]:


ax1 = data.iloc[0:5,:].plot(kind = 'bar',figsize = (20,15), fontsize = 30, rot = 0)
ax1.set_ylabel('consecutive cases / year',fontdict={'fontsize':35}, labelpad = 20)
ax1.set_xlabel('no. of consecutive fog events',fontdict={'fontsize':35}, labelpad= 20)
ax1.legend(loc=1,fontsize=25)
ax1.set_yticks(range(0,21,1))
ax1.set_yticklabels(['0','' ,'','','','5','','','','', '10','' ,'','','',
                    '15','' ,'','','','20'])

#xticks(range(1, 5))

ax1.yaxis.set_ticks_position('both')
ax1.xaxis.set_ticks_position('both')
ax1.tick_params(axis='x', which='major', pad=15)
ax1.tick_params(axis='y', which='major', pad=15)
#left, bottom, width, height = [.30, 0.6, 0.2, 0.25]
#ax_new = ax1.([left, bottom, width, height])



# In[84]:


bx = data.iloc[5:20,:].plot(kind = 'bar',figsize = (20,15), fontsize = 30, rot =0)
bx.set_xlabel("no. of consecutive fog events", fontdict={'fontsize':35}, labelpad=20)
bx.set_ylabel("conseucutive cases / year", fontdict = {'fontsize':35}, labelpad = 20)
bx.legend(loc=1,fontsize=25)

bx.set_yticks(list(np.arange(0.0,0.6,0.1, dtype = '<f')))
#bx.set_yticklabels(['0','' ,'','','','5','','','','', '10','' ,'','','',
 #                   '15','' ,'','','','20'])
    
    

bx.yaxis.set_ticks_position('both')
bx.xaxis.set_ticks_position('both')

bx.tick_params(axis='x', which='major', pad=10)
bx.tick_params(axis='y', which='major', pad=10)


# In[87]:


### visualizing monthly distribution of foggy events for the 6 stations 
### ((Budapest, Zagreb, Nagyszeben, Ungvar, Pecs, Hereny))
data_2 = {'Budapest': [10.52,5.81,1.32,0.45,0.16,0.16,0.06,0.35,0.68,3.81,7.42,10.84], 
        'Pécs':       [12.26,6.85,2.15,0.37,0.19,0.15, 0.11,0.30,0.7,4.26,11.44,14.81],
        'Nagyszeben':[9.81, 5.32, 1.29, 1.06, 0.00, 0.06, 0.19,0.16, 0.39, 0.97, 5.19, 11.65],
        'Ungvár':[9.86, 5.25, 1.82, 0.25, 0, 0, 0.04, 0, 0.36, 1.64, 5.86, 12.00],
        'Herény':[6.67,3.67, 1.15, 0.41, 0.19, 0.11, 0.19, 0.0, 0.44, 3.48, 4.48, 6.59],
        'Zágráb':[9, 4.31, 0.97, 0.21, 0.07, 0.07, 0.24, 0.38, 1.48,5.28, 7.52, 12.55]
       }


# In[88]:


data_2 = pd.DataFrame(data_2)
Months = ['Január', 'Február', 'Március', 'Április', 'Május', 'Június',
         'Július', 'Augusztus', 'Szeptember', 'Október', 'November', 'December']

Months_2 = ['January', 'February', 'March', 'April', 'May', 'June',
         'July', 'August', 'September', 'October', 'November', 'December']


# In[89]:


data_2.index = Months_2
cx = data_2.plot(kind = 'bar',figsize = (20,15), fontsize = 30,rot = 45)
cx.set_xlabel("Months", fontdict={'fontsize':35},labelpad= -10)
cx.set_ylabel("cases / year", fontdict = {'fontsize':35}, labelpad = 10)
cx.legend(loc=0,fontsize=25)


cx.yaxis.set_ticks_position('both')
cx.xaxis.set_ticks_position('both')
cx.tick_params(axis='x', which='major', pad=10)
cx.tick_params(axis='y', which='major', pad=10)


# In[91]:


#cx.figure.savefig('/Volumes/Seagate Backup Plus Drive/historical/plot3.png')
plt.show()